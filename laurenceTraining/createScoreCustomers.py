import openpyxl
import configparser
import requests


# Read Oracle.ini file and get variables
parser = configparser.ConfigParser()
parser.read('init.ini')
token_url=parser['SCORE']['token_url']

def getscoretoken():
    response = requests.get(token_url)
    data = response.json()
    token = data["sessionKey"]
    return (token)

def main():
    #call function getscoretoken
    token = getscoretoken()
    print(token)


    wb = openpyxl.load_workbook('scoreCustomers.xlsx')
    ws = wb['Sheet1']
    print(ws.max_row)

    for i in range(1, ws.max_row+1):
        print("Row ", i, " data :")

        for j in range(1, ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            print(cell_obj.value, end=" ")

# {
#  "cusExRef":		"CLTEST0002",
#  "cusName": 		"CL TEST 002",
#  "cusAdr1": 		"avenue prince de liege 74",
#  "cusCity": 		"Jambes",
#  "cusCtry": 		"BEL",
#  "cusPosCode":  	"5100",
#  "cusPhones":
# [
# {"cusPhoneNbr": 	"1235896",
#  "cusPhoneType":	"DOM"}
# ],
# "cusLangCode":  	"NL",
# "cusCurr":  		"EUR",
# "cusPayCurr":   	"EUR",
# "cusBankAcc":   	"BE42063449267154",
# "cusBankAccBankCode":"TT",
# "cusContact":		"BNH"
# }


if __name__ == '__main__' :
     main()