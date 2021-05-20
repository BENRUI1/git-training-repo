# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import datetime
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

import openpyxl

wb = openpyxl.load_workbook('scoreCustomers.xlsx')
ws = wb['Sheet1']
print(ws.max_row)

for i in range(1, ws.max_row+1):
    print("\n")
    print("Row ", i, " data :")

    for j in range(1, ws.max_column+1):
        cell_obj = ws.cell(row=i, column=j)
        print(cell_obj.value, end=" ")